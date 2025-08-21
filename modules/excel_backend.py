from openpyxl import Workbook, load_workbook
import os


def create_excel_file(path):
    wb = Workbook()
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # Créer les feuilles dans un ordre précis
    wb.create_sheet('Étudiants', 0)
    wb.create_sheet('Paiements', 1)
    wb.create_sheet('Suivi', 2)

    # Optionnel : ajouter en-têtes dans les feuilles (exemple minimal)
    etudiants_sheet = wb['Étudiants']
    etudiants_sheet.append(['Nom', 'Prénom', 'Matricule'])  # Exemple colonnes

    paiements_sheet = wb['Paiements']
    paiements_sheet.append(['Étudiant', 'Montant', 'Date'])

    suivi_sheet = wb['Suivi']
    suivi_sheet.append(['Date', 'Événement', 'Commentaire'])

    # Création du répertoire si nécessaire
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb.save(path)


def load_workbook_for_classe(classe_name):
    filepath = f"data/classes/{classe_name}.xlsx"
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Le fichier pour la classe '{classe_name}' n'existe pas : {filepath}")
    try:
        return load_workbook(filepath)
    except Exception as e:
        raise IOError(f"Erreur lors du chargement du fichier {filepath} : {e}")
